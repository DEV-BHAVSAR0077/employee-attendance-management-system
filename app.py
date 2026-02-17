from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import os
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import sqlite3
from backend.gemini_service import GeminiReportService
from backend.pdf_generator import AttendanceReportPDF
from backend.time_utils import calculate_working_hours, calculate_status, format_hours
from backend.db_utils import init_db, get_settings
from backend.excel_utils import parse_excel_file

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database setup
# Database initialization moved to db_utils

# Initialize database on startup
init_db()

@app.route('/')
def index():
    """Render home page"""
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    """Render dashboard page"""
    return render_template('dashboard.html')

@app.route('/reports')
def reports():
    """Render reports page"""
    return render_template('reports.html')

@app.route('/employees')
def employees():
    """Render employees page"""
    return render_template('employees.html')

@app.route('/settings')
def settings():
    """Render settings page"""
    return render_template('settings.html')

@app.route('/ai-reports')
def ai_reports_page():
    """Render AI-powered reports page"""
    return render_template('ai_reports.html')

@app.route('/api/upload', methods=['POST'])
def upload_attendance():
    """Upload and process attendance Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        target_date = request.form.get('target_date')  # Get optional target date

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Only Excel files are allowed'}), 400
        
        # Save file first to parse it
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name, ext = os.path.splitext(filename)
        unique_filename = f"{base_name}_{timestamp}{ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # Parse Excel file to get dates
        try:
            records = parse_excel_file(filepath)
        except Exception as e:
            # If parsing fails, remove the file and error out
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'success': False, 'error': f"Failed to parse file: {str(e)}"}), 400

        if not records:
             if os.path.exists(filepath):
                os.remove(filepath)
             return jsonify({'success': False, 'error': 'Uploaded file contains no records.'}), 400

        # Determine date from records
        # If multiple dates exist, we pick the first one or min one as the "primary" date for validaiton
        file_dates = sorted(list(set(r['date'] for r in records)))
        primary_date = file_dates[0] if file_dates else None

        # Logic: If target_date is provided (from calendar), validation matches that.
        # If not provided (generic upload), we use the date found in the file.
        effective_date = target_date if target_date else primary_date
        
        if not effective_date:
             # Should not happen if records exist
             return jsonify({'success': False, 'error': 'Could not determine date from file.'}), 400

        # Validation on effective_date
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        # 1. Future check
        today_str = datetime.now().strftime('%Y-%m-%d')
        if effective_date > today_str:
            conn.close()
            return jsonify({'success': False, 'error': f'Cannot upload data for future date ({effective_date}).'}), 400

        # 2. Duplicate check (One upload per date)
        c.execute("SELECT id FROM upload_history WHERE target_date = ?", (effective_date,))
        if c.fetchone() is not None:
             conn.close()
             return jsonify({'success': False, 'error': f'Data for {effective_date} has already been uploaded. You can only upload one file per date.'}), 400
             
        # Optional: Check if file content dates match target_date if target_date WAS provided
        if target_date and primary_date != target_date:
             if target_date not in file_dates:
                  conn.close()
                  return jsonify({'success': False, 'error': f'Selected date is {target_date}, but file contains data for {file_dates}. Please upload the correct file.'}), 400

        # If we passed validation, preserve the effective_date as the target_date for history
        final_target_date = effective_date
        
        # Determine date range for response
        min_date = min(file_dates) if file_dates else None
        max_date = max(file_dates) if file_dates else None
        
        records_processed = len(records)
        records_success = 0
        records_failed = 0
        
        # Insert records into database
        for record in records:
            try:
                # Check if employee exists, create if not
                c.execute('''INSERT OR IGNORE INTO employees 
                                (employee_id, employee_name, is_active) 
                                VALUES (?, ?, 1)''', 
                                (record['employee_id'], record['employee_name']))
                
                c.execute('''INSERT OR REPLACE INTO attendance_records 
                             (employee_id, employee_name, date, punch_in_time, punch_out_time, 
                              working_hours, status, month, year, break_start_time, break_end_time,
                              break_duration, is_late, break_exceeded, is_break_outside_window, is_early_departure)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                          (record['employee_id'], record['employee_name'], record['date'], 
                           record['punch_in_time'], record['punch_out_time'], record['working_hours'], 
                           record['status'], record['month'], record['year'],
                           record.get('break_start_time'), record.get('break_end_time'),
                           record.get('break_duration'), record.get('is_late'),
                           record.get('break_exceeded'), record.get('is_break_outside_window'), record.get('is_early_departure')))
                records_success += 1
            except Exception as e:
                print(f"Error inserting record: {e}")
                records_failed += 1
        
        # Record upload history with target_date
        c.execute('''INSERT INTO upload_history 
                     (file_name, file_path, records_processed, records_success, records_failed, status, target_date)
                     VALUES (?, ?, ?, ?, ?, ?, ?)''',
                  (filename, filepath, records_processed, records_success, records_failed, 'success', final_target_date))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'File processed successfully',
            'recordsProcessed': records_processed,
            'recordsSuccess': records_success,
            'recordsFailed': records_failed,
            'maxDate': max_date,
            'filePath': filepath,
            'fileName': filename
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/attendance/range')
def get_attendance_by_range():
    """Get attendance records by date range"""
    try:
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        employee_id = request.args.get('employeeId')
        
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        query = '''SELECT * FROM attendance_records
                  WHERE date BETWEEN ? AND ?'''
        params = [start_date, end_date]
        
        if employee_id:
            query += ' AND employee_id = ?'
            params.append(employee_id)
        
        query += ' ORDER BY date DESC'
        
        c.execute(query, params)
        columns = [description[0] for description in c.description]
        records = [dict(zip(columns, row)) for row in c.fetchall()]
        
        conn.close()
        
        return jsonify({'records': records})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/attendance/date')
def get_attendance_by_date():
    """Get attendance records for a specific date"""
    try:
        date = request.args.get('date')
        
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        c.execute('''SELECT * FROM attendance_records
                    WHERE date = ?
                    ORDER BY employee_name''', (date,))
        
        columns = [description[0] for description in c.description]
        records = [dict(zip(columns, row)) for row in c.fetchall()]
        
        conn.close()
        
        return jsonify({'records': records})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/employees')
def get_employees():
    """Get all employees"""
    try:
        employee_id = request.args.get('employeeId')
        
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        if employee_id:
            c.execute('SELECT * FROM employees WHERE employee_id = ? ORDER BY employee_name', (employee_id,))
        else:
            c.execute('SELECT * FROM employees ORDER BY employee_name')
        
        columns = [description[0] for description in c.description]
        employees = [dict(zip(columns, row)) for row in c.fetchall()]
        
        conn.close()
        
        return jsonify({'employees': employees})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload-history')
def get_upload_history():
    """Get upload history"""
    try:
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        c.execute('''SELECT * FROM upload_history
                    ORDER BY upload_date DESC
                    LIMIT 50''')
        
        columns = [description[0] for description in c.description]
        history = [dict(zip(columns, row)) for row in c.fetchall()]
        
        conn.close()
        
        return jsonify({'history': history})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics')
def get_statistics():
    """Get attendance statistics with optional filtering"""
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        employee_id = request.args.get('employeeId')
        
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        # Build filter conditions
        query_conditions = []
        params = []
        
        if month and year:
            query_conditions.append('month = ? AND year = ?')
            params.extend([month, year])
        
        if start_date and end_date:
            query_conditions.append('date BETWEEN ? AND ?')
            params.extend([start_date, end_date])
        elif start_date:
            query_conditions.append('date >= ?')
            params.append(start_date)
        elif end_date:
            query_conditions.append('date <= ?')
            params.append(end_date)
            
        if employee_id:
            query_conditions.append('employee_id = ?')
            params.append(employee_id)
        
        where_clause = 'WHERE ' + ' AND '.join(query_conditions) if query_conditions else ''
        
        # Total records
        c.execute(f'SELECT COUNT(*) FROM attendance_records {where_clause}', params)
        total_records = c.fetchone()[0]
        
        # Present count
        present_params = params + ["Present"]
        c.execute(f'SELECT COUNT(*) FROM attendance_records {where_clause} {"AND" if where_clause else "WHERE"} status = ?',
                 present_params)
        present_count = c.fetchone()[0]
        
        # Absent count
        absent_params = params + ["Absent"]
        c.execute(f'SELECT COUNT(*) FROM attendance_records {where_clause} {"AND" if where_clause else "WHERE"} status = ?',
                 absent_params)
        absent_count = c.fetchone()[0]
        
        # Average working hours
        hours_params = params
        c.execute(f'SELECT AVG(working_hours) FROM attendance_records {where_clause} {"AND" if where_clause else "WHERE"} working_hours IS NOT NULL',
                 hours_params)
        avg_hours = c.fetchone()[0] or 0
        
        # Total employees
        # If we are filtering by a specific employee, the total employees is 1 (if they have records in range)
        if employee_id:
            c.execute(f'SELECT COUNT(DISTINCT employee_id) FROM attendance_records {where_clause}', params)
        else:
            # If no employee filter, show unique employees in the filtered records
            if where_clause:
                c.execute(f'SELECT COUNT(DISTINCT employee_id) FROM attendance_records {where_clause}', params)
            else:
                # Fallback to total active employees if no filters at all
                c.execute('SELECT COUNT(DISTINCT employee_id) FROM employees')
        
        total_employees = c.fetchone()[0]
        
        conn.close()
        
        return jsonify({
            'totalRecords': total_records,
            'presentCount': present_count,
            'absentCount': absent_count,
            'averageWorkingHours': round(avg_hours, 2),
            'totalEmployees': total_employees,
            'attendanceRate': round((present_count / total_records * 100) if total_records > 0 else 0, 2)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete-upload', methods=['DELETE'])
def delete_upload():
    """Delete all uploaded attendance data"""
    try:
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        # Delete all attendance records
        c.execute('DELETE FROM attendance_records')
        
        # Delete upload history
        c.execute('DELETE FROM upload_history')
        
        # Delete employees (optional but cleaner for full reset)
        c.execute('DELETE FROM employees')
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'All attendance data deleted successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download-latest-file')
def download_latest_file():
    """Download the most recently uploaded file"""
    try:
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        # Get the most recent upload with file path
        c.execute('''SELECT file_name, file_path FROM upload_history 
                    WHERE file_path IS NOT NULL 
                    ORDER BY upload_date DESC 
                    LIMIT 1''')
        
        result = c.fetchone()
        conn.close()
        
        if not result:
            return jsonify({'success': False, 'error': 'No uploaded file found'}), 404
        
        original_filename, file_path = result
        

        # Check if file exists
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': 'File not found on server'}), 404
        
        # Send file for download
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=original_filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/calendar-dates')
def get_calendar_dates():
    """Get all dates that have attendance data"""
    try:
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        # Get distinct dates from attendance_records
        c.execute('SELECT DISTINCT date FROM attendance_records ORDER BY date')
        dates = [row[0] for row in c.fetchall()]
        
        conn.close()
        
        return jsonify({'dates': dates})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel')
def export_excel():
    """Export attendance records to Excel"""
    try:
        month = request.args.get('month', type=int)
        year = request.args.get('year', type=int)
        
        conn = sqlite3.connect('attendance.db')
        c = conn.cursor()
        
        query = 'SELECT * FROM attendance_records'
        params = []
        
        if month and year:
            query += ' WHERE month = ? AND year = ?'
            params.extend([month, year])
        
        query += ' ORDER BY date DESC, employee_name'
        
        c.execute(query, params)
        columns = [description[0] for description in c.description]
        records = c.fetchall()
        conn.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {month}-{year}" if month and year else "Attendance Report"
        
        # Header style
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = ['Employee ID', 'Employee Name', 'Date', 'Punch In', 'Punch Out', 'Working Hours', 'Break Start', 'Break End', 'Break Duration (Mins)', 'Late?', 'Break Exceeded?', 'Off Schedule Break?', 'Early Departure?', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        settings = get_settings()
        for row_idx, record in enumerate(records, 2):
            # record structure: 0:id, 1:emp_id, 2:emp_name, 3:date, 4:punch_in, 5:punch_out, 6:work_hours, 7:status, 8:month, 9:year, 11:break_start, 12:break_end, 13:break_dur
            
            # Helper to get column value safely (in case of old records without new columns)
            def get_col(idx, default=None):
                try:
                    return record[idx]
                except IndexError:
                    return default

            ws.cell(row=row_idx, column=1, value=record[1])  # employee_id
            ws.cell(row=row_idx, column=2, value=record[2])  # employee_name
            ws.cell(row=row_idx, column=3, value=record[3])  # date
            ws.cell(row=row_idx, column=4, value=record[4])  # punch_in_time
            ws.cell(row=row_idx, column=5, value=record[5])  # punch_out_time
            ws.cell(row=row_idx, column=6, value=format_hours(record[6]))  # working_hours
            
            break_start = get_col(11) # Assuming column index 11
            break_end = get_col(12)   # Assuming column index 12
            
            ws.cell(row=row_idx, column=7, value=break_start)
            ws.cell(row=row_idx, column=8, value=break_end)
            
            # Recalculate status and flags on-the-fly
            status_result = calculate_status(record[4], record[5], record[6], record[3], settings, break_start, break_end)
            
            ws.cell(row=row_idx, column=9, value=status_result['break_duration'])
            ws.cell(row=row_idx, column=10, value="Yes" if status_result['is_late'] else "No")
            ws.cell(row=row_idx, column=11, value="Yes" if status_result['break_exceeded'] else "No")
            ws.cell(row=row_idx, column=12, value="Yes" if status_result['is_break_outside_window'] else "No")
            ws.cell(row=row_idx, column=13, value="Yes" if status_result['is_early_departure'] else "No")
            
            ws.cell(row=row_idx, column=14, value=status_result['status'])
            
            # Highlight Late/Break Exceeded/Off Schedule/Early Departure
            if status_result['is_late']:
                 ws.cell(row=row_idx, column=10).font = Font(color="FF0000", bold=True)
            if status_result['break_exceeded']:
                 ws.cell(row=row_idx, column=11).font = Font(color="FF0000", bold=True)
            if status_result['is_break_outside_window']:
                 ws.cell(row=row_idx, column=12).font = Font(color="FF0000", bold=True)
            if status_result['is_early_departure']:
                 ws.cell(row=row_idx, column=13).font = Font(color="FF0000", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"attendance_{month}_{year}.xlsx" if month and year else "attendance_report.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/settings', methods=['GET', 'POST'])
def data_settings():
    """Get or update specific settings"""
    conn = sqlite3.connect('attendance.db')
    c = conn.cursor()
    
    if request.method == 'POST':
        try:
            data = request.get_json()
            for key, value in data.items():
                c.execute('INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)', (key, str(value)))
            conn.commit()
            return jsonify({'success': True, 'message': 'Settings updated successfully'})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
        finally:
            conn.close()
    else:
        try:
            c.execute('SELECT key, value FROM settings')
            settings = dict(c.fetchall())
            
            # Ensure defaults for required keys
            defaults = {
                'standard_start_time': '09:30',
                'max_break_duration': '60'
            }
            
            return jsonify({**defaults, **settings})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            conn.close()

@app.route('/api/recalculate', methods=['POST'])
def recalculate_status_route():
    """Recalculate status for all attendance records based on current settings"""
    try:
        conn = sqlite3.connect('attendance.db')
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        
        # specific_date = request.json.get('date') # Optional: if we want to target specific date
        
        c.execute('SELECT * FROM attendance_records')
        rows = c.fetchall()
        
        updated_count = 0
        settings = get_settings()
        
        for row in rows:
            # Recalculate
            status_result = calculate_status(
                row['punch_in_time'], 
                row['punch_out_time'], 
                row['working_hours'], 
                row['date'], 
                settings,
                row['break_start_time'], 
                row['break_end_time']
            )
            
            # Update DB
            c.execute('''UPDATE attendance_records SET 
                         status = ?, 
                         break_duration = ?, 
                         is_late = ?, 
                         break_exceeded = ?, 
                         is_break_outside_window = ?, 
                         is_early_departure = ?
                         WHERE id = ?''',
                      (status_result['status'], 
                       status_result['break_duration'], 
                       1 if status_result['is_late'] else 0, 
                       1 if status_result['break_exceeded'] else 0, 
                       1 if status_result['is_break_outside_window'] else 0, 
                       1 if status_result['is_early_departure'] else 0,
                       row['id']))
            updated_count += 1
            
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': f'Updated {updated_count} records with new settings.'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============= GEMINI AI REPORT GENERATION ROUTES =============

@app.route('/api/gemini/configure', methods=['POST'])
def configure_gemini():
    """Configure Gemini API key"""
    try:
        data = request.get_json()
        api_key = data.get('apiKey')
        
        if not api_key:
            return jsonify({'error': 'API key is required'}), 400
        
        # Save API key to environment or config file
        os.environ['GEMINI_API_KEY'] = api_key
        
        # Test the API key
        try:
            service = GeminiReportService(api_key=api_key)
            if service.model:
                return jsonify({
                    'success': True,
                    'message': 'Gemini API key configured successfully'
                })
            else:
                return jsonify({'error': 'Invalid API key'}), 400
        except Exception as e:
            return jsonify({'error': f'API key validation failed: {str(e)}'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gemini/usage')
def get_gemini_usage():
    """Get Gemini API usage statistics"""
    try:
        service = GeminiReportService()
        usage_stats = service.get_usage_stats()
        
        # Estimate free tier limits (Gemini 1.5 Flash free tier)
        # Free tier: 15 requests/minute, 1500 requests/day, 1M tokens/day
        daily_limit = 1000000  # 1M tokens per day
        usage_percentage = (usage_stats['today_tokens'] / daily_limit) * 100
        
        return jsonify({
            'today_tokens': usage_stats['today_tokens'],
            'month_tokens': usage_stats['month_tokens'],
            'today_cached': usage_stats['today_cached'],
            'today_new': usage_stats['today_new'],
            'daily_limit': daily_limit,
            'usage_percentage': round(usage_percentage, 2),
            'tokens_remaining': daily_limit - usage_stats['today_tokens']
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gemini/generate-report', methods=['POST'])
def generate_ai_report():
    """Generate AI-powered attendance report"""
    try:
        data = request.get_json()
        report_type = data.get('reportType', 'daily')
        start_date = data.get('startDate')
        end_date = data.get('endDate')
        force_regenerate = data.get('forceRegenerate', False)
        
        # Initialize Gemini service
        service = GeminiReportService()
        
        # Check if API key is configured
        if not service.model:
            return jsonify({
                'error': 'Gemini API not configured',
                'message': 'Please configure your Gemini API key in settings'
            }), 400
        
        # Generate report
        report_data = service.generate_report(
            db_path='attendance.db',
            report_type=report_type,
            start_date=start_date,
            end_date=end_date,
            force_regenerate=force_regenerate
        )
        
        if 'error' in report_data:
            return jsonify(report_data), 500
        
        return jsonify({
            'success': True,
            'report': report_data,
            'cached': report_data.get('cached', False)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/gemini/download-pdf', methods=['POST'])
def download_ai_report_pdf():
    """Generate and download PDF report"""
    try:
        data = request.get_json()
        report_type = data.get('reportType', 'custom')
        start_date = data.get('startDate')
        end_date = data.get('endDate')
        
        # Generate report data
        service = GeminiReportService()
        
        if not service.model:
            return jsonify({
                'error': 'Gemini API key not configured',
                'message': 'Please add GEMINI_API_KEY to your .env file or environment variables'
            }), 400
        
        report_data = service.generate_report(
            db_path='attendance.db',
            report_type=report_type,
            start_date=start_date,
            end_date=end_date
        )
        
        if 'error' in report_data:
            return jsonify(report_data), 500
        
        # Add report type to data
        report_data['report_type'] = report_type.capitalize()
        
        # Generate PDF
        pdf_generator = AttendanceReportPDF()
        pdf_buffer = pdf_generator.generate_pdf(report_data)
        
        # Create filename
        date_suffix = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"attendance_ai_report_{report_type}_{date_suffix}.pdf"
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'message': 'Failed to generate AI report. Check server logs for details.'
        }), 500

@app.route('/api/gemini/clear-cache', methods=['POST'])
def clear_report_cache():
    """Clear cached reports"""
    try:
        import shutil
        cache_dir = 'report_cache'
        
        if os.path.exists(cache_dir):
            shutil.rmtree(cache_dir)
            os.makedirs(cache_dir, exist_ok=True)
        
        return jsonify({
            'success': True,
            'message': 'Report cache cleared successfully'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
